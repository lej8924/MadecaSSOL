from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
#{{import 선배 파일에 있는 클래스 혹은 변수들}}

wb = Workbook()

ws = wb.create_sheet('estimate')

# ws.title="estimate"
#####################

align_center = Alignment(horizontal='center', vertical='center')
very_center = Alignment(horizontal="center", vertical="center")

fill_lightgray = PatternFill('solid', fgColor='ededed')
fill_gray = PatternFill('solid', fgColor='bdbebd')
fill_thickgray = PatternFill('solid', fgColor='737a7c')

font_bold = Font(size=8.5, bold=True, color='000000')
font_small = Font(size=8.5,name="Nanum Gothic")

top_border = Border(top=Side(border_style='thin', color='000000'))
bottom_border = Border(bottom=Side(border_style='thin', color='000000'))
dot_bottom_border = Border(bottom=Side(border_style='dotted', color='000000'))
dot_top_border = Border(top=Side(border_style='dotted', color='000000'))


#####################

for i in range(12):
    for k in range(35):
        ws[chr(65+i)+str(8+k)].font = font_small
        
for i in range(15):
    tmp1 = ws['F'+str(9+i)]
    tmp1.fill  = fill_thickgray
    tmp2 = ws['H'+str(9+i)]
    tmp2.fill  = fill_thickgray
    tmp3 = ws['J'+str(9+i)]
    tmp3.fill  = fill_thickgray

#####################
ws.merge_cells('A1:C4')
ws['A1'] = "견적서"
ws['A1'].font = Font(color="000000", size=22, bold=True)
ws['A1'].alignment = align_center

for i in range(4):
    for k in range(4):
        ws[chr(68+i)+str(1+k)].border = dot_bottom_border
        
        
for i in range(4):
    for k in range(4):
        ws[chr(73+i)+str(1+k)].border = dot_bottom_border
        

ws['D1'] = "견적명"
ws['D1'].font = font_bold
ws.merge_cells('E1:G1')


ws['D2'] = "발주자"
ws['D2'].font = font_bold
ws.merge_cells('E2:G2')


ws['D3'] = "견적일"
ws['D3'].font = font_bold
ws.merge_cells('E3:G3')


ws['D4'] = "유효기간"
ws['D4'].font = font_bold
ws.merge_cells('E4:G4')


ws['I1'] = "공급자 상호"
ws['I1'].font = font_bold
ws['J1'] = ""

ws['I2'] ="등록번호"
ws['I2'].font = font_bold
ws.merge_cells('J2:L2')
ws['J2'] = ""

ws['I3'] = "주소"
ws['I3'].font = font_bold
ws.merge_cells('J3:L3')

ws['I4'] ="전화"
ws['I4'].font = font_bold
ws['J4'] = ""

ws['K1'] = "대표자"
ws['K1'].font = font_bold
ws['L1'] = "(인)"

ws['K4'] ="팩스"
ws['K4'].font= font_bold
ws['L4'] = ""

ws['A6'] = "사업면적"
ws['A6'].font = Font(size=14, bold=True, color='ff0000',name="Nanum Gothic")


ws['B6'] = "헥타르"
ws['B6'].font = Font(size=14, bold=True, color='ff0000',name="Nanum Gothic")

ws['C6'] = "ha"
ws['C6'].font = Font(size=14, bold=True, color='ff0000',name="Nanum Gothic")
for i in range(3):
    tmp = ws[chr(65+i)+"6"]
    tmp.fill = fill_lightgray

ws.merge_cells('D6:F6')
ws['D6'] = "견적금액(공급가액+세액)"
ws['D6'].font = Font(size=14, bold=True, name="Nanum Gothic")

ws.row_dimensions[6].height = 24
ws.row_dimensions[7].height = 5

##################################################################3
for i in range(9):
    tmp = ws[chr(68+i)+"6"]
    tmp.fill = fill_lightgray



for i in range(12):
    tmp = ws[chr(65+i)+"8"]
    tmp.fill = fill_gray
    tmp.border = top_border


name =["구분","항목","세부항목","","인력등급","월 노임단가","1일 노임단가","투입기간(일)","투입인원/장비수","참여율(%)","금액","소계"]
for i in range(12):
    ws[chr(65+i)+"8"] = name[i]
    ws[chr(65+i)+"8"].font = font_bold


for i in range(12):
    tmp = ws[chr(65+i)+"24"]
    tmp.fill = fill_gray
    tmp.border = Border(bottom=Side(border_style='thin', color='000000'),top=Side(border_style="dotted",color='000000'))

ws.merge_cells('A9:A14')
ws['A9'] = "직접인건비"
ws['A9'].alignment = very_center

for i in range(12):
    tmp = ws[chr(65+i)+"15"]
    tmp.border = Border(bottom=Side(border_style='dotted', color='000000'),top=Side(border_style="dotted",color='000000'))

ws['B9'] = "설계서 검토"

ws['B10'] = "현장지도"

ws.merge_cells('B11:B12')
ws['B11'] = "사업평가"
ws['B11'].alignment = very_center

ws['E11'] = "특급기술자"
ws['E12'] = "초급기술자"

ws['B13'] = "사업비정산"
ws['B14'] = "감리보고서 작성"

ws['A15'] = "직접경비"
ws['B15'] = "재료유인비"





ws.merge_cells('A16:A21')
ws['A16'] = "소나무단가"
ws['A16'].alignment = very_center

pinetree = ["특용재급","1등급","2등급","3등급","원주재급","원료재급"]
for i in range(6):
    ws['B'+str(16+i)] = pinetree[i]

for i in range(12):
    tmp = ws[chr(65+i)+"22"]
    tmp.border = Border(top=Side(border_style="dotted",color='000000'))


ws.merge_cells('A22:A23')
ws['A22'] = "운반비"
ws['A22'].alignment = very_center
ws['B22'] = "5톤"
ws['B23'] = "25톤"



ws['A24'] = "소계"
ws['A24'].font = font_bold

ws['A25'] = "직접경비"
ws.merge_cells('B25:K25')
ws['B25'] = "사진 이미지 구입, 식비, 교통비(별도 견적 참조)"

ws['A26'] = "제경비"
ws.merge_cells('B26:K26')
ws['B26'] ="직접인건비*110%"

ws['A27'] = "기술료"
ws.merge_cells('B27:K27')
ws['B27']="(직접인건비+제경비) * 20%"


ws['A28']="합계"
ws['A28'].font = font_bold
ws.merge_cells('B28:E28')
ws['B28'] = "직접인건비+직접경비+제경비+기술료"
ws['B28'].font = font_bold


for i in range(12):
    tmp = ws[chr(65+i)+"28"]
    tmp.fill = fill_gray
    tmp.border =Border(bottom=Side(border_style='thin', color='000000'),top=Side(border_style="dotted",color='000000'))

ws['A30']="기타"

for i in range(12):
    tmp = ws[chr(65+i)+"30"]
    tmp.fill = fill_gray
    tmp.border = top_border
######################
ws.merge_cells('L9:L14') 
ws['L9']= "=SUM(K9:K14)"

ws['L15'] = "=K15"

ws.merge_cells('L16:L21')
ws['L16']= "=SUM(K16:K21)"

ws.merge_cells("L22:L23")
ws['L22']= "=SUM(K22:K23)"

ws['L25'] = "=L15"
ws['L26'] = "=L9*110%"
ws['L27'] = "=(L9+L26)*20%"

#####################
##선배 파일 경로 설정해서 저장!
wb.save("c:/madecassol/test03.xlsx")
wb.close()